# -*- coding: utf-8 -*-
"""
Created on Thu May 22 17:03:11 2025

@author: yogesh.sanjay.gavade
"""

# === src/open/prb.py ===
import time
import pandas as pd
from openpyxl import load_workbook
from config import OPEN_INC_FILE, METRICS_FILE
from src.utils.excel_utils import (
    copy_filtered_rows,
    build_lookup_dict,
    apply_quarterly_sla_filter
)

def run():
    start_time = time.time()

    # Step 1: Copy filtered PRB rows to 'Open PRBs'
    copy_filtered_rows(
        source_file=OPEN_INC_FILE,
        target_file=METRICS_FILE,
        target_sheet="Open PRBs",
        starts_with="PRB"
    )

    # Step 2: Apply VLOOKUP-like mappings
    wb = load_workbook(METRICS_FILE)
    sheet = wb['Open PRBs']
    ag_sheet = wb['AG to Portfolio mapping']

    portfolio_dict = build_lookup_dict(ag_sheet, match_col=1, value_col=2)
    qn_dict = build_lookup_dict(ag_sheet, match_col=1, value_col=4)

    for row in sheet.iter_rows(min_row=2):
        key = str(row[1].value).strip().lower()
        row[21].value = portfolio_dict.get(key, "#N/A")
        row[20].value = qn_dict.get(key, "#N/A")

    wb.save(METRICS_FILE)

    print("✅ VLOOKUPs applied for Portfolio and Q/N in Open PRBs.")

    # Step 3: Filter data in pandas
    df = pd.read_excel(METRICS_FILE, sheet_name="Open PRBs")
    qn_sla_pairs = [
        ('Quarterly', 'Accenture_P4 Defect Closure (Quarterly)_App Dev SLA'),
        ('Non-Quterly', 'Accenture_P4 Defect Closure (Non-Quarterly)_App Dev SLA')
    ]

    final_df = df.copy()
    for qn_value, sla in qn_sla_pairs:
        final_df = apply_quarterly_sla_filter(final_df, qn_value, sla)

    # Step 4: Add formula columns
    final_df['SLA Bucket'] = [
        f'=IF(K{row}<=25,"Within 25% SLA time",'
        f'IF(AND(K{row}>25,K{row}<=50),"Reached 25-50% SLA time",'
        f'IF(AND(K{row}>50,K{row}<=75),"Reached 50-75% SLA time",'
        f'IF(AND(K{row}>75,K{row}<100),"Reached 75-100% SLA time","Missed SLA"))))'
        for row in range(2, len(final_df) + 2)
    ]

    final_df['Business elapsed time (Days)'] = [f'=J{row}/60/60/24' for row in range(2, len(final_df) + 2)]

    with pd.ExcelWriter(METRICS_FILE, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        final_df.to_excel(writer, sheet_name='Open PRBs', index=False)

    elapsed_time = time.time() - start_time
    print(f"✅ Open PRB module completed in {elapsed_time:.2f} seconds.")