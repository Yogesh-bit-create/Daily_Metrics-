# -*- coding: utf-8 -*-
"""
Created on Thu May 22 16:50:50 2025

@author: yogesh.sanjay.gavade
"""

# === src/open/incident.py ===
import time
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import warnings
from config import OPEN_INC_FILE, METRICS_FILE, OPEN_INC_AGE_FILE

from src.utils.excel_utils import (
    filter_assignment_groups,
    filter_states,
    copy_filtered_rows,
    apply_priority_sla_filter,
    build_lookup_dict,
    apply_days_awaiting_expiration
)

def run():
    start_time = time.time()
    warnings.filterwarnings("ignore", message="Workbook contains no default style")

    current_date = datetime.now().date()

    # Step 1: Clean open incident file
    wb = load_workbook(OPEN_INC_FILE)
    sheet = wb.active
    sheet.auto_filter.ref = None

    filter_assignment_groups(sheet, ["IT.A.PAS-Help_Desk", "IT.A.PAS-Triage"])
    filter_states(sheet, ["Closed", "Resolved"])
    wb.save(OPEN_INC_FILE)
    print("✅ Cleaned assignment group and state filters in Open INC file.")

    # Step 2: Copy filtered INC rows
    copy_filtered_rows(
        source_file=OPEN_INC_FILE,
        target_file=METRICS_FILE,
        target_sheet="Open INCs",
        starts_with="INC"
    )

    # Step 3: SLA filtering
    df = pd.read_excel(METRICS_FILE, sheet_name="Open INCs")
    priority_sla_pairs = [
        ('3 - Medium', 'Accenture_P3 Incident Resolution_App Dev SLA'),
        ('4 - Low', 'Accenture_P4 Incident Resolution_App Dev SLA'),
        ('5 - Minimal', 'Accenture_P4 Incident Resolution_App Dev SLA'),
        ('2 - High', 'Accenture_P2 Incident Resolution_App Dev SLA')
    ]

    final_df = df.copy()
    for priority, sla in priority_sla_pairs:
        final_df = apply_priority_sla_filter(final_df, priority, sla)

    # Remove duplicate rows marked "Completed"
    final_df = final_df[~((final_df.duplicated(subset='Task', keep=False)) & (final_df['Stage'] == 'Completed'))]

    # Step 4: Add SLA Buckets and calculated columns
    final_df['SLA Bucket'] = [
        f'=IF(K{row}<=25,"Within 25% SLA time",'
        f'IF(AND(K{row}>25,K{row}<=50),"Reached 25-50% SLA time",'
        f'IF(AND(K{row}>50,K{row}<=75),"Reached 50-75% SLA time",'
        f'IF(AND(K{row}>75,K{row}<100),"Reached 75-100% SLA time","Missed SLA"))))'
        for row in range(2, len(final_df) + 2)
    ]

    final_df['Business elapsed time (Days)'] = [f'=J{row}/60/60/24' for row in range(2, len(final_df) + 2)]
    final_df['Ageing Bucket (Days)'] = [
        f'=IF(W{row}>100,">100 Days",IF(W{row}>=50,"50-100 Days",IF(W{row}>=30,"30-49 Days",IF(W{row}>=22,"22-29 Days","<22 Days"))))'
        for row in range(2, len(final_df) + 2)
    ]

    with pd.ExcelWriter(METRICS_FILE, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        final_df.to_excel(writer, sheet_name='Open INCs', index=False)

    # Step 5: Post-save enhancements
    workbook = load_workbook(METRICS_FILE)
    sheet = workbook['Open INCs']
    ag_map_sheet = workbook['AG to Portfolio mapping']
    open_age_wb = load_workbook(OPEN_INC_AGE_FILE)
    open_age_sheet = open_age_wb.active

    sheet['AB1'] = current_date

    # Portfolio mapping (column 21)
    portfolio_lookup = build_lookup_dict(ag_map_sheet, match_col=1, value_col=2)
    for row in sheet.iter_rows(min_row=2):
        key = str(row[1].value).strip().lower()
        row[20].value = portfolio_lookup.get(key, "#N/A")

    # Q/N lookup from ageing file (columns 25 & 26)
    qn_dict1 = build_lookup_dict(open_age_sheet, match_col=1, value_col=10)
    qn_dict2 = build_lookup_dict(open_age_sheet, match_col=1, value_col=11)

    for row in sheet.iter_rows(min_row=2):
        key = str(row[2].value).strip().lower()
        row[24].value = qn_dict1.get(key, "#N/A")
        row[25].value = qn_dict2.get(key, "#N/A")

    apply_days_awaiting_expiration(sheet)
    workbook.save(METRICS_FILE)

    elapsed_time = time.time() - start_time
    print(f"✅ Open Incident module completed in {elapsed_time:.2f} seconds.")
