# -*- coding: utf-8 -*-
"""
Created on Fri May 23 00:22:27 2025

@author: yogesh.sanjay.gavade
"""

# === src/prb_categorization.py ===
import openpyxl
from config import METRICS_FILE

SOURCE_FILE = r"C:\Excel\Defects Delivered for Daily Reporting.xlsx"

def run():
    # Load workbooks
    source_wb = openpyxl.load_workbook(SOURCE_FILE)
    source_ws = source_wb.active

    dest_wb = openpyxl.load_workbook(METRICS_FILE)
    ag_map_sheet = dest_wb['AG to Portfolio mapping']

    if "PRB Categorization base" in dest_wb.sheetnames:
        dest_ws = dest_wb['PRB Categorization base']
    else:
        dest_ws = dest_wb.create_sheet("PRB Categorization base")

    # Copy rows from source to destination starting from row 2
    for idx, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True), start=2):
        for col_idx, value in enumerate(row, start=1):
            dest_ws.cell(row=idx, column=col_idx, value=value)

    # Build portfolio lookup
    lookup_dict = {
        str(row[1]).strip().lower(): row[2]
        for row in ag_map_sheet.iter_rows(min_row=1, max_col=3, values_only=True)
    }

    # Apply VLOOKUP logic to column 21 using column 16 as key
    for row in dest_ws.iter_rows(min_row=2):
        key = str(row[15].value).strip().lower()
        row[20].value = lookup_dict.get(key, "#NA")

    dest_wb.save(METRICS_FILE)

    print("✅ PRB Categorization base data updated and portfolio lookup applied.")
