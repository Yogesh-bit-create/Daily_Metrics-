# -*- coding: utf-8 -*-
"""
Created on Fri May 23 00:27:25 2025

@author: yogesh.sanjay.gavade
"""

# === src/last_updated_incident.py ===
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from config import METRICS_FILE

SOURCE_FILE = r"C:\Excel\Last Updated Incident Report.xlsx"

def run():
    start_time = datetime.now()
    current_date = start_time.date()

    df = pd.read_excel(SOURCE_FILE)
    ag_portfolio = pd.read_excel(METRICS_FILE, sheet_name='AG to Portfolio mapping')
    open_incs = pd.read_excel(METRICS_FILE, sheet_name='Open INCs')

    # Build portfolio lookup
    portfolio_dict = {
        str(row[1]).strip().lower(): row[2]
        for _, row in ag_portfolio.iterrows()
    }

    # Build Has Task SLA lookup from Open INCs
    task_sla_dict = {
        str(row[2]).strip(): row[2]
        for _, row in open_incs.iterrows()
    }

    # Filter out PAS assignment groups
    exclude_groups = ["IT.A.PAS-Help_Desk", "IT.A.PAS-Triage"]
    filtered_df = df[~df['Assignment group'].isin(exclude_groups)].copy()

    # Add calculated and lookup fields
    filtered_df['Portfolio'] = [
        portfolio_dict.get(str(row[10]).strip().lower(), "#N/A")
        for _, row in filtered_df.iterrows()
    ]

    filtered_df['Days since last Updated'] = [f'=$AD$1-V{row}' for row in range(2, len(filtered_df) + 2)]

    filtered_df['Last Updated Ageing'] = [
        f'=IF(X{row}>50,"> 50 Days",IF(X{row}>30,"> 30 Days",IF(X{row}>20,"> 20 Days",IF(X{row}>14,"> 14 Days",IF(X{row}>7,"> 7 Days","< 7 Days")))))'
        for row in range(2, len(filtered_df) + 2)
    ]

    filtered_df['Open Incident Ageing (Days)'] = [f'=$AD$1-N{row}' for row in range(2, len(filtered_df) + 2)]

    filtered_df['Open Incident Ageing'] = [
        f'=IF(Z{row}>300,"> 300 Days",IF(Z{row}>200,"> 200 Days",IF(Z{row}>100,"> 100 Days",IF(Z{row}>75,"> 75 Days",'
        f'IF(Z{row}>60,"> 60 Days",IF(Z{row}>50,"> 50 Days",IF(Z{row}>30,"30-50 Days",IF(Z{row}>22,"22-30 Days","< 22 Days"))))))))'
        for row in range(2, len(filtered_df) + 2)
    ]

    filtered_df['Assigned?'] = [f'=IF(L{row}<>"","Yes","No")' for row in range(2, len(filtered_df) + 2)]

    filtered_df['Has Task SLA?'] = [
        task_sla_dict.get(str(row[1]).strip(), "#N/A") for _, row in filtered_df.iterrows()
    ]

    # Save to Excel
    with pd.ExcelWriter(METRICS_FILE, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        filtered_df.to_excel(writer, sheet_name='Open Incidents last Updated', index=False)

    wb = load_workbook(METRICS_FILE)
    ws = wb['Open Incidents last Updated']
    ws['AD1'] = current_date
    wb.save(METRICS_FILE)

    elapsed_time = (datetime.now() - start_time).total_seconds()
    print(f"✅ Last Updated Incident module completed in {elapsed_time:.2f} seconds.")
