# -*- coding: utf-8 -*-
"""
Created on Thu May 22 16:33:31 2025

@author: yogesh.sanjay.gavade
"""
print("this is my first code change on git")

print("hey this is new feature")

# === src/closed/incident.py ===
import time
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import warnings

from config import CLOSED_INC_FILE, METRICS_FILE
from src.utils.excel_utils import (
    filter_assignment_groups,
    filter_states,
    filter_by_date_column,
    copy_filtered_rows,
    apply_priority_sla_filter,
    retain_best_duplicate,
    apply_sla_formula,
    apply_portfolio_lookup
)

def run():
    start_time = time.time()
    warnings.filterwarnings("ignore", message="Workbook contains no default style")

    # Load workbook
    wb = load_workbook(CLOSED_INC_FILE)
    ws = wb.active
    ws.auto_filter.ref = None

    # Clean data
    filter_assignment_groups(ws, ["IT.A.PAS-Help_Desk", "IT.A.PAS-Triage", "WWO.CNA.Agency_Help_Desk"])
    filter_states(ws, ["Cancelled"])
    current_month = datetime.now().month
    current_year = datetime.now().year
    filter_by_date_column(ws, "Stop time", current_month, current_year)
    filter_by_date_column(ws, "Resolved", current_month, current_year)

    # Save intermediate cleaned file
    wb.save(CLOSED_INC_FILE)

    # Copy filtered rows to 'Closed INCs' sheet
    copy_filtered_rows(
        source_file=CLOSED_INC_FILE,
        target_file=METRICS_FILE,
        target_sheet="Closed INCs",
        starts_with="INC"
    )

    # Load into pandas and apply SLA filters
    df = pd.read_excel(METRICS_FILE, sheet_name="Closed INCs")
    priority_sla_pairs = [
        ('3 - Medium', 'Accenture_P3 Incident Resolution_App Dev SLA'),
        ('4 - Low', 'Accenture_P4 Incident Resolution_App Dev SLA'),
        ('5 - Minimal', 'Accenture_P4 Incident Resolution_App Dev SLA'),
        ('2 - High', 'Accenture_P2 Incident Resolution_App Dev SLA')
    ]
    final_df = df.copy()
    for priority, sla in priority_sla_pairs:
        final_df = apply_priority_sla_filter(final_df, priority, sla)

    with pd.ExcelWriter(METRICS_FILE, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        final_df.to_excel(writer, sheet_name='Closed INCs', index=False)

    # Reload updated workbook to apply post-processing
    wb = load_workbook(METRICS_FILE)
    sheet = wb['Closed INCs']
    sheet2 = wb['AG to Portfolio mapping']

    retain_best_duplicate(sheet, check_col='Task', compare_col='Business elapsed time')
    apply_sla_formula(sheet, sla_col="K", result_col_name="SLA Status")
    apply_portfolio_lookup(sheet, sheet2, target_col=21, match_col_index=1)

    wb.save(METRICS_FILE)
    elapsed_time = time.time() - start_time
    print(f"Closed Incident module executed in {elapsed_time:.2f} seconds.")
