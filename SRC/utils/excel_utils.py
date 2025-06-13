from openpyxl import load_workbook
import pandas as pd

def filter_assignment_groups(ws, values_to_delete):
    header = {cell.value: cell.column for cell in ws[1]}
    idx = header.get("Assignment group")
    if not idx:
        raise ValueError("Assignment group column not found.")
    rows = [row[0].row for row in ws.iter_rows(min_row=2) if row[idx - 1].value in values_to_delete]
    for row_idx in reversed(rows):
        ws.delete_rows(row_idx)

def filter_states(ws, values_to_delete):
    header = {cell.value: cell.column for cell in ws[1]}
    idx = header.get("State")
    if not idx:
        raise ValueError("State column not found.")
    rows = [row[0].row for row in ws.iter_rows(min_row=2) if row[idx - 1].value in values_to_delete]
    for row_idx in reversed(rows):
        ws.delete_rows(row_idx)

def filter_by_date_column(ws, column_header, month, year):
    idx = None
    for cell in ws[1]:
        if cell.value == column_header:
            idx = cell.column
            break
    if not idx:
        raise ValueError(f"{column_header} not found.")
    rows = [row[0].row for row in ws.iter_rows(min_row=2)
            if hasattr(row[idx - 1].value, 'month') and
               (row[idx - 1].value.month != month or row[idx - 1].value.year != year)]
    for row_idx in reversed(rows):
        ws.delete_rows(row_idx)

def copy_filtered_rows(source_file, target_file, target_sheet, starts_with="INC"):
    src_wb = load_workbook(source_file)
    src_ws = src_wb.active
    tgt_wb = load_workbook(target_file)
    tgt_ws = tgt_wb[target_sheet]
    header = {cell.value: cell.column for cell in src_ws[1]}
    task_idx = header.get("Task")
    if not task_idx:
        return
    row_idx = 2
    for row in src_ws.iter_rows(min_row=2, max_col=src_ws.max_column):
        val = row[task_idx - 1].value
        if isinstance(val, str) and val.startswith(starts_with):
            for col_idx, cell in enumerate(row, 1):
                tgt_ws.cell(row=row_idx, column=col_idx).value = cell.value
            row_idx += 1
    tgt_wb.save(target_file)

def apply_priority_sla_filter(df, priority, sla_def):
    duplicates = df[df.duplicated(subset=['Task'], keep=False)]
    priority_df = duplicates[duplicates['Priority'] == priority]
    non_match = priority_df[priority_df['SLA definition'] != sla_def]
    return df.drop(non_match.index)

def apply_quarterly_sla_filter(df, qn_value, sla_def):
    duplicates = df[df.duplicated(subset=['Task'], keep=False)]
    qn_df = duplicates[duplicates['Quarterly / Non-Quarterly P4 SLA designation'] == qn_value]
    non_matching = qn_df[qn_df['SLA definition'] != sla_def]
    return df.drop(non_matching.index)

def retain_best_duplicate(sheet, check_col, compare_col):
    headers = [cell.value for cell in sheet[1]]
    check_idx = headers.index(check_col) + 1
    compare_idx = headers.index(compare_col) + 1
    best_rows = {}
    for row in sheet.iter_rows(min_row=2):
        key = row[check_idx - 1].value
        val = row[compare_idx - 1].value
        if key in best_rows:
            if val > best_rows[key][1]:
                best_rows[key] = (row, val)
        else:
            best_rows[key] = (row, val)
    rows_to_delete = [row[0].row for row in sheet.iter_rows(min_row=2)
                      if (row[check_idx - 1].value in best_rows and
                          (row, row[compare_idx - 1].value) != best_rows[row[check_idx - 1].value])]
    for idx in sorted(set(rows_to_delete), reverse=True):
        sheet.delete_rows(idx)

def apply_sla_formula(sheet, sla_col, result_col_name):
    headers = [cell.value for cell in sheet[1]]
    if result_col_name not in headers:
        return
    col_idx = headers.index(result_col_name) + 1
    for i in range(2, sheet.max_row + 1):
        sheet.cell(row=i, column=col_idx).value = f'=IF({sla_col}{i}<100,"Met","Not Met")'

def apply_portfolio_lookup(sheet, mapping_sheet, target_col, match_col_index, value_col=2):
    lookup = {}
    for row in mapping_sheet.iter_rows(min_row=1, values_only=True):
        try:
            key = str(row[match_col_index]).strip().lower()
            value = row[value_col]
            lookup[key] = value
        except Exception:
            continue
    for row in sheet.iter_rows(min_row=2):
        try:
            key = str(row[1].value).strip().lower()
            row[target_col - 1].value = lookup.get(key, "#N/A")
        except Exception:
            row[target_col - 1].value = "#N/A"

def build_lookup_dict(sheet, match_col=1, value_col=2):
    return {
        str(row[match_col]).strip().lower(): row[value_col]
        for row in sheet.iter_rows(min_row=1, values_only=True)
        if row[match_col] is not None
    }

def apply_days_awaiting_expiration(sheet, column_name="Days Awaiting expiration", date_cell="AB1"):
    col = None
    for c in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=c).value == column_name:
            col = c
            break
    if col:
        for r in range(2, sheet.max_row + 1):
            sheet.cell(row=r, column=col).value = f'=IF(Y{r}="",0,Y{r}-$${date_cell}$$)'

